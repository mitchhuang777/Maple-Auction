import pyautogui
import time
from cnocr import CnOcr
from PIL import Image, ImageGrab
import threading
import openpyxl
import datetime
import os

start_time = time.time()
game_x, game_y, game_w, game_h = pyautogui.locateOnScreen('maple-icon.png')
pyautogui.moveTo(game_x + game_w//2, game_y + game_h//2, duration=0.2)
pyautogui.click()
time.sleep(1)

filename = 'maple-auto-auction.xlsx'


try:
    consume_pos = pyautogui.locateOnScreen('consume.png')
    goto_pos = pyautogui.center(consume_pos)
    pyautogui.moveTo(goto_pos, duration=0.2)
    pyautogui.click()
except:
    print("No consume button")

try:
    consume_item_all_pos = pyautogui.locateOnScreen('consume-item-all.png')
    goto_pos = pyautogui.center(consume_item_all_pos)
    pyautogui.moveTo(goto_pos, duration=0.2)
    pyautogui.click()

    monster_pos = pyautogui.locateOnScreen('monster.png')
    goto_pos = pyautogui.center(monster_pos)
    pyautogui.moveTo(goto_pos, duration=0.2)
    pyautogui.click()
except:
    print("No consume item all button")

try:
    search_all_pos = pyautogui.locateOnScreen('search-all.png')
    goto_pos = pyautogui.center(search_all_pos)
    pyautogui.moveTo(goto_pos, duration=0.2)
    pyautogui.click()

    monster_pos = pyautogui.locateOnScreen('search-all-monster.png')
    goto_pos = pyautogui.center(monster_pos)
    pyautogui.moveTo(goto_pos, duration=0.2)
    pyautogui.click()
except:
    print("No search all button")

start_search_pos = pyautogui.locateOnScreen('start-search.png')
goto_pos = pyautogui.center(start_search_pos)
pyautogui.moveTo(goto_pos, duration=0.2)
pyautogui.click()

confirm_pos = pyautogui.locateOnScreen('confirm.png')
goto_pos = pyautogui.center(confirm_pos)
pyautogui.moveTo(goto_pos, duration=0.2)
pyautogui.click()

confirm_pos = pyautogui.locateOnScreen('confirm-2.png')
goto_pos = pyautogui.center(confirm_pos)
pyautogui.moveTo(goto_pos, duration=0.2)
pyautogui.click()


while(True):
    try:
        warn_pos = pyautogui.locateOnScreen('warning.png')
        goto_pos = pyautogui.center(warn_pos)
        pyautogui.moveTo(goto_pos, duration=0.2)
        pyautogui.click()
    except:
        pass
        
    next_pos = pyautogui.locateOnScreen('next.png')
    goto_pos = pyautogui.center(next_pos)
    pyautogui.moveTo(goto_pos, duration=0.2)
    pyautogui.click()
    
    try:
        confirm_pos = pyautogui.locateOnScreen('confirm-3-1.png') 
        goto_pos = pyautogui.center(confirm_pos)
        pyautogui.moveTo(goto_pos, duration=0.2)
        pyautogui.click()
    except:
        pass
    try:    
        confirm_pos = pyautogui.locateOnScreen('confirm-3-2.png')
        goto_pos = pyautogui.center(confirm_pos)
        pyautogui.moveTo(goto_pos, duration=0.2)
        pyautogui.click()
    except:
        pass
    
    # if find the end-next button, then break
    try:
        next_pos = pyautogui.locateOnScreen('end-next.png')
        if next_pos:
            break
    except:
        pass
    time.sleep(5)

price_sort_pos = pyautogui.locateOnScreen('price-sort.png')
goto_pos = pyautogui.center(price_sort_pos)
pyautogui.moveTo(goto_pos, duration=0.2)
pyautogui.click()
time.sleep(0.5)
pyautogui.click()

ocr_naive_det = CnOcr(det_model_name='naive_det')
ocr_en_pp = CnOcr(det_model_name='en_PP-OCRv3_det', rec_model_name='en_PP-OCRv3')

card_types_1 = {
    'legend.png': '傳說',
    'legend-5.png': '傳說-5',
    'legend-2.png': '傳說-2',
    'legend-4.png': '傳說-4',
    'legend-3.png': '傳說-3'

    
}


card_types_2 = {
    'special.png': '特殊',
    'rare.png': '稀有',
    'epic.png': '罕見'
}


card_images_1 = set(card_types_1.keys())
card_images_2 = set(card_types_2.keys())

def get_card_type_1(card_img, card_types, flag):
    for image_name in card_images_1:
        if pyautogui.locateOnScreen(image_name) is not None:
            card_type = card_types.get(image_name)
            # print(f"Card type: {card_type}")
            card_types_found.append(card_type)
            flag = True
            return 
    # card_types_found.append('None')


def get_card_type_2(card_img, card_types, flag):
    for image_name in card_images_2:
        if pyautogui.locateOnScreen(image_name) is not None:
            card_type = card_types.get(image_name)
            # print(f"Card type: {card_type}")
            card_types_found.append(card_type)
            flag = True
            return 

'''
workbook = openpyxl.Workbook()
worksheet = workbook.active
'''
# page = 1
end_of_flag = False
if __name__ == "__main__":
    try:
    # Open xlsx file, the filename is maple-auto-auction.xlsx
        workbook = openpyxl.load_workbook('maple-auto-auction.xlsx')
    except:
        workbook = openpyxl.Workbook()
        
    worksheet = workbook.active
    
    
    search_result_pos = pyautogui.locateOnScreen('search-result.png')
    goto_pos = pyautogui.center(search_result_pos)
    game_x, game_y = goto_pos
    screenshot = ImageGrab.grab(bbox=(game_x+336, game_y-12, game_x+336+28, game_y-12+25))
    img = Image.frombytes('RGB', screenshot.size, screenshot.tobytes())
    page_result = ocr_en_pp.ocr(img)
    total_page = page_result[0]['text']
    
    search_result_pos = pyautogui.locateOnScreen('search-result.png')
    goto_pos = pyautogui.center(search_result_pos)
    game_x, game_y = goto_pos
    screenshot = ImageGrab.grab(bbox=(game_x+280, game_y-12, game_x+300+28, game_y-12+25))
    img = Image.frombytes('RGB', screenshot.size, screenshot.tobytes())
    page_result = ocr_en_pp.ocr(img)
    img.save('test.png')
    try:
        page = int(page_result[0]['text'])
    except:
        page = 1
    
    while(True):
    # for i in range(0, 3):
        count = 0
        search_result_pos = pyautogui.locateOnScreen('search-result.png')
        goto_pos = pyautogui.center(search_result_pos)
        # goto_pos x座標 - 10
        goto_pos = (goto_pos[0], goto_pos[1])
        pyautogui.moveTo(goto_pos, duration=0.2)
        
        # Card List
        game_x, game_y = goto_pos
        screenshot = ImageGrab.grab(bbox=(game_x+35, game_y+45, game_x+35+240, game_y+45+493))
        img = Image.frombytes('RGB', screenshot.size, screenshot.tobytes())

        card_result = ocr_naive_det.ocr(img)
        
        # Write the result to the excel      
        for i in range(0, len(card_result)):
            worksheet.cell(row=i+1+(page-1)*9, column=1).value = card_result[i]['text']
            # worksheet.cell(row=i+1+(page-1)*9, column=4).value = card_result[i]['score']
        
        # Price List
        game_x, game_y = goto_pos
        screenshot = ImageGrab.grab(bbox=(game_x+255, game_y+45, game_x+255+170, game_y+45+493))
        img = Image.frombytes('RGB', screenshot.size, screenshot.tobytes())
        price_result = ocr_en_pp.ocr(img)
        
        for i in range(0, 18, 2):
            # 取得文字
            # try break the loop if IndexError: list index out of range
            try:
                price_string = price_result[i]['text']
                price_string = price_string.replace('Q', '0')  # replace 'Q' with '0'
                price_string = price_string.replace(',', '')  # remove commas
                price_string = price_string.replace('.', '')  # remove dots
                price_string = price_string.replace(' ', '')  # remove spaces
                value = int(price_string)
                worksheet.cell(row=i//2+1+(page-1)*9, column=2).value = value
            except:
                pass
                
        pyautogui.moveRel(0, 68, duration=0.001)
        card_result = [{'text': 'legend.png'}]
        card_types_found = []
        flag = False
        t1 = threading.Thread(target=get_card_type_1, args=(card_result[0]['text'], card_types_1, flag))
        t2 = threading.Thread(target=get_card_type_2, args=(card_result[0]['text'], card_types_2, flag))

        t1.start()
        t2.start()
        
        t1.join()
        
        if flag == False:
            t2.join()
        try:
            worksheet.cell(row=1+(page-1)*9, column=3).value = card_types_found[-1]
        except:
            worksheet.cell(row=1+(page-1)*9, column=3).value = 'None'
        
        locate_ability_pos = pyautogui.locateOnScreen('locate-ability.png')
        goto_pos = pyautogui.center(locate_ability_pos)
        game_x, game_y = goto_pos
        screenshot = ImageGrab.grab(bbox=(game_x, game_y+48, game_x+230, game_y+48+96))
        img = Image.frombytes('RGB', screenshot.size, screenshot.tobytes())
        card_ability = ocr_naive_det.ocr(img)
        
        for i in range(1, len(card_ability)):
            worksheet.cell(row=1+(page-1)*9, column=i+5).value = card_ability[i]['text']
        
        '''
        for i in range(1, len(card_ability)):
            worksheet.cell(row=1+(page-1)*9, column=i+11).value = card_ability[i]['score']
        '''
        
        # print(len(price_result))
        for i in range(0, (len(price_result)//2)-1):
            flag = False
            pyautogui.moveRel(0, 55, duration=0.001)
            end_block_pos = pyautogui.locateOnScreen('end-block.png')
            if end_block_pos is None:
                count += 1
            
            card_type = None
            t1 = threading.Thread(target=get_card_type_1, args=(card_result[0]['text'], card_types_1, flag))
            t2 = threading.Thread(target=get_card_type_2, args=(card_result[0]['text'], card_types_2, flag))

            t1.start()
            t2.start()
            
            t1.join()
            
            if flag == False:
                t2.join()
            
            worksheet.cell(row=i+2+(page-1)*9, column=3).value = card_types_found[-1]
            
            locate_ability_pos = pyautogui.locateOnScreen('locate-ability.png')
            goto_pos = pyautogui.center(locate_ability_pos)
            game_x, game_y = goto_pos
            screenshot = ImageGrab.grab(bbox=(game_x, game_y+48, game_x+230, game_y+48+96))
            img = Image.frombytes('RGB', screenshot.size, screenshot.tobytes()) 
            card_ability = ocr_naive_det.ocr(img)
            
            for j in range(0, len(card_ability)):
                worksheet.cell(row=2+i+(page-1)*9, column=j+5).value = card_ability[j]['text']
            
            '''
            for j in range(0, len(card_ability)):
                worksheet.cell(row=2+i+(page-1)*9, column=j+11).value = card_ability[j]['score']
            '''
            
        next_page_pos = pyautogui.locateOnScreen('next-page.png')
        goto_pos = pyautogui.center(next_page_pos)
        print(f"Page {page} done")
        pyautogui.moveTo(goto_pos, duration=0.001)
        pyautogui.click()
        if int(page) == int(total_page):
            break
        
        page += 1
        
        card_types_found = []
        if end_of_flag == True:
            break
            
        def seconds_to_hms(seconds):
            """將秒數轉換為時:分:秒的格式"""
            hours, remainder = divmod(seconds, 3600)
            minutes, seconds = divmod(remainder, 60)
            return f"{int(hours):02d}:{int(minutes):02d}:{int(seconds):02d}"
        
        workbook.save(filename)
        
        end_time = time.time()
        elapsed_time = end_time - start_time
        hms_time = seconds_to_hms(elapsed_time)
        print(f"Elapsed time: {hms_time}")
        
        
    
    '''
    timestamp = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
    filename = f"{timestamp}-auction.xlsx"
    workbook.save(filename)
    def seconds_to_hms(seconds):
        """將秒數轉換為時:分:秒的格式"""
        hours, remainder = divmod(seconds, 3600)
        minutes, seconds = divmod(remainder, 60)
        return f"{int(hours):02d}:{int(minutes):02d}:{int(seconds):02d}"
    
    end_time = time.time()
    elapsed_time = end_time - start_time
    hms_time = seconds_to_hms(elapsed_time)
    print(f"Elapsed time: {hms_time}")
    print("Done")
    '''