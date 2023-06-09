import openpyxl
import re
import time

# 開啟 Excel 檔案
wb = openpyxl.load_workbook('maple-auto-auction.xlsx')
# wb = openpyxl.load_workbook('123-before.xlsx')

# 選擇要操作的工作表
ws = wb['Sheet']

# 取得工作表的最大列數和最大欄數
max_row = ws.max_row
max_col = ws.max_column

# 設定進度條的起始值和結束值
start_value = 0
end_value = max_row * max_col

# 計算目前進度條的值
def get_progress_value(row, col):
    return (row - 1) * max_col + col

# 迭代每一個儲存格，找出特定字串並取代
for row in range(1, max_row + 1):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        if isinstance(cell.value, str):
            cell.value = re.sub(r'\s+', '', cell.value)
            cell.value = cell.value.replace('加持技能持技时周', '加持技能持續時間')
            cell.value = cell.value.replace('加持技能持纩时周', '加持技能持續時間')
            cell.value = cell.value.replace('加持技能持绩时周', '加持技能持續時間')
            cell.value = cell.value.replace('跳暹力', '跳躍力')
            cell.value = cell.value.replace('跳耀力', '跳躍力')
            cell.value = cell.value.replace('魔法攻车力', '魔法攻擊力')
            cell.value = cell.value.replace('魔法攻季力', '魔法攻擊力')
            cell.value = cell.value.replace('全厨性', '全屬性')
            cell.value = cell.value.replace('全网性', '全屬性')
            cell.value = cell.value.replace('防票力', '防禦力')
            cell.value = cell.value.replace('最终僖害', '最終傷害')
            cell.value = cell.value.replace('最终但害', '最終傷害')
            cell.value = cell.value.replace('黑视怪物防票车', '無視怪物防禦力')
            cell.value = cell.value.replace('集视怪物防票车', '無視怪物防禦力')
            cell.value = cell.value.replace('慧视怪物防票车', '無視怪物防禦力')
            cell.value = cell.value.replace('移勤速度', '移動速度')
            cell.value = cell.value.replace('物理攻车力', '物理攻擊力')
            cell.value = cell.value.replace('物理攻季力', '物理攻擊力')
            cell.value = cell.value.replace('物理攻辜力', '物理攻擊力')
            cell.value = cell.value.replace('4秒内恢便', '4秒內恢復')
            cell.value = cell.value.replace('4秒内恢复', '4秒內恢復')
            cell.value = cell.value.replace('被勤技能', '被動技能')
            cell.value = cell.value.replace('斑法攻车力', '魔法攻擊力')
            cell.value = cell.value.replace('埋车摄车', '爆擊機率')
            cell.value = cell.value.replace('埕车横车', '爆擊機率')
            cell.value = cell.value.replace('埋车提车', '爆擊機率')
            cell.value = cell.value.replace('埕车摄车', '爆擊機率')
            cell.value = cell.value.replace('车力', '擊力')
            cell.value = cell.value.replace('攻车时有125摄车蛋勤3级的中毒效', '攻擊時有12%機率發動3級的中毒效')
            cell.value = cell.value.replace('攻车时有10%摄车蛋勤3级的中毒效', '攻擊時有10%機率發動3級的中毒效')
            cell.value = cell.value.replace('攻车时有125摄车蛋勤3级的冰结效', '攻擊時有12%機率發動3級的冰結效')
            cell.value = cell.value.replace('攻车时有12%摄车蛋勤3级的冰结效', '攻擊時有12%機率發動3級的冰結效')
            cell.value = cell.value.replace('攻车时有10%摄车蛋勤3级的冰结效', '攻擊時有10%機率發動3級的冰結效')
            cell.value = cell.value.replace('攻车时有125摄车蛋勤3级的封印效', '攻擊時有12%機率發動3級的封印效')
            cell.value = cell.value.replace('攻车时有12%摄车蛋勤3级的封印效', '攻擊時有12%機率發動3級的封印效')
            cell.value = cell.value.replace('攻车时有10%摄车蛋勤3级的封印效', '攻擊時有10%機率發動3級的封印效')
            cell.value = cell.value.replace('攻季时有5%摄车蛋勤2级的封印效果', '攻擊時有5%機率發動2級的封印效果')
            cell.value = cell.value.replace('攻车时有125摄车蛋勤3级的缓慢效', '攻擊時有12%機率發動3級的緩慢效')
            cell.value = cell.value.replace('攻车时有125摄车蛋勤3级的量该效', '攻擊時有12%機率發動3級的暈眩效')
            cell.value = cell.value.replace('攻车时有125摄车蛋勤3级的盟黑效', '攻擊時有12%機率發動3級的暗黑效')
            cell.value = cell.value.replace('攻车时有12%摄车蛋勤3级的阁黑效', '攻擊時有12%機率發動3級的暗黑效')
            cell.value = cell.value.replace('攻车时有10%摄车蛋勤3级的阁黑效', '攻擊時有10%機率發動3級的暗黑效')
            cell.value = cell.value.replace('攻车时有3%摄车恢复', '攻擊時有3%機率恢復')
            cell.value = cell.value.replace('攻车时有10%摄车蛋勤3级的量该效', '攻擊時有10%機率發動3級的暈眩效')
            cell.value = cell.value.replace('攻车时有10%摄车蛋勤3级的量眩效', '攻擊時有10%機率發動3級的暈眩效')
            cell.value = cell.value.replace('攻车时有5%横车劲勤3级的量眩效果', '攻擊時有5%機率發動3級的暈眩效果')
            cell.value = cell.value.replace('攻车时有75横车劲勤3级的经慢效果', '攻擊時有7%機率發動3級的緩慢效果')
            cell.value = cell.value.replace('攻车时有10%摄车蛋勤3级的缓慢效', '攻擊時有10%機率發動3級的緩慢效')
            cell.value = cell.value.replace('攻车时有12%摄车蛋勤3级的缓慢效', '攻擊時有12%機率發動3級的緩慢效')
            cell.value = cell.value.replace('攻季时有5%摄车蛋勤1级的冰结效果', '攻擊時有5%機率發動1級的冰結效果')
            cell.value = cell.value.replace('攻季时有5%摄车蛋勤1级的量眩效果', '攻擊時有5%機率發動1級的暈眩效果')
            cell.value = cell.value.replace('攻季时有5%摄车蛋勤1级的量眩效果', '攻擊時有5%機率發動1級的暈眩效果')
            cell.value = cell.value.replace('攻季时有5%摄车蛋勤3级的量眩效果', '攻擊時有5%機率發動3級的暈眩效果')
            cell.value = cell.value.replace('攻车时有12%摄车蛋勤3级的量眩效', '攻擊時有12%機率發動3級的暈眩效')
            cell.value = cell.value.replace('依照角色全部网性的90%来追加萌的', '依照角色全部屬性的90%來追加萌獸的')
            cell.value = cell.value.replace('依照角色全部网性的70%来追加萌的', '依照角色全部屬性的70%來追加萌獸的')
            cell.value = cell.value.replace('依照角色全部网性的20%来追加萌的', '依照角色全部屬性的20%來追加萌獸的')
            cell.value = cell.value.replace('依照角色攻季力的90%来追加萌鲜的攻', '依照角色攻擊力的90%來追加萌獸的攻')
            cell.value = cell.value.replace('依照角色攻季力的70%来追加萌鲜的攻', '依照角色攻擊力的70%來追加萌獸的攻')
            cell.value = cell.value.replace('攻季力', '攻擊力')
            cell.value = cell.value.replace('季力', '擊力')
            cell.value = cell.value.replace('攻车时有3%摄车恢馍', '攻擊時有3%機率恢復')
            cell.value = cell.value.replace('攻车时有3%摄车恢馍', '攻擊時有3%機率恢復')
            cell.value = cell.value.replace('攻车时有3%横车恢复', '攻擊時有3%機率恢復')
            cell.value = cell.value.replace('攻车时有3%摄车恢便', '攻擊時有3%機率恢復')
            cell.value = cell.value.replace('攻车时有5%摄车蛋勤2级的冰结效果', '攻擊時有5%機率發動2級的冰結效果')
            cell.value = cell.value.replace('攻车时有5%摄车蛋勤2级的封印效果', '攻擊時有5%機率發動2級的封印效果')
            cell.value = cell.value.replace('攻车时有5%摄车蛋勤2级的盟黑效果', '攻擊時有5%機率發動2級的暗黑效果')
            cell.value = cell.value.replace('攻车时有5%摄车蛋勤2级的封印效果', '攻擊時有5%機率發動2級的封印效果')
            cell.value = cell.value.replace('攻车时有5%摄车蛋勤3级的冰结效果', '攻擊時有5%機率發動3級的冰結效果')
            cell.value = cell.value.replace('攻车时有5%摄车蛋勤3级的盟黑效果', '攻擊時有5%機率發動3級的暗黑效果')
            cell.value = cell.value.replace('攻季时有3%横车恢馍', '攻擊時有3%機率恢復')
            cell.value = cell.value.replace('攻车时有3%横车恢馍', '攻擊時有3%機率恢復')
            cell.value = cell.value.replace('依照角色全部网性的85%来追加萌的', '依照角色全部屬性的85%來追加萌獸的')
            cell.value = cell.value.replace('依照角色全部网性的80%来追加萌的', '依照角色全部屬性的80%來追加萌獸的')
            cell.value = cell.value.replace('依照角色全部国性的20%来追加萌的', '依照角色全部屬性的20%來追加萌獸的')
            cell.value = cell.value.replace('攻车时有12%摄车蛋勤3级的中毒效', '攻擊時有12%機率發動3級的中毒效')
            cell.value = cell.value.replace('攻车时有12%摄车蛋勤3级的量该效', '攻擊時有12%機率發動3級的暈眩效')
            cell.value = cell.value.replace('攻车时有5%摄车蛋勤3级的中毒效果', '攻擊時有5%機率發動3級的中毒效果')
            cell.value = cell.value.replace('攻季时有3%横车恢复', '攻擊時有3%機率恢復')
            cell.value = cell.value.replace('攻季力','攻擊力')
            cell.value = cell.value.replace('攻车力','攻擊力')
            cell.value = cell.value.replace('僖害', '傷害')
            
            cell.value = cell.value.replace('土莆卡', '土龍卡')
            cell.value = cell.value.replace('大副凯丁卡', '大副凱丁卡')
            cell.value = cell.value.replace('小小萌默', '小小萌獸')
            cell.value = cell.value.replace('小幻门影使盗卡', '小幻影俠盜卡')
            cell.value = cell.value.replace('小企鹅王娃娃栈卡', '小企鵝王娃娃機卡')
            cell.value = cell.value.replace('小拉尼王卡', '小拉尼亞卡')
            cell.value = cell.value.replace('小障月卡', '小隱月卡')
            cell.value = cell.value.replace('小融卡', '小蘭卡')
            cell.value = cell.value.replace('山妖萌默', '山妖萌獸')
            cell.value = cell.value.replace('马目立克α萌默', '馬貝立克α萌獸')
            cell.value = cell.value.replace('不协调精需卡', '不協調精靈卡')
            cell.value = cell.value.replace('之幼红蜀角狮卡', '幼紅獨角獅卡')
            cell.value = cell.value.replace('凶暴的莫沙萌默', '凶暴的莫沙萌獸')
            cell.value = cell.value.replace('天生童身的麻烦裂造者萌蟹', '天生單身的麻煩製造者萌獸')
            cell.value = cell.value.replace('巴逵雷伊萌默', '巴達雷伊萌獸')
            cell.value = cell.value.replace('水祭司萌默', '水祭司萌獸')
            cell.value = cell.value.replace('水精需卡', '水精靈卡')
            cell.value = cell.value.replace('火牢衔卡', '火牢術卡')
            cell.value = cell.value.replace('火焰艾南逵斯卡', '火焰艾爾達斯卡')
            cell.value = cell.value.replace('火焰祭司萌默', '火焰祭司萌獸')
            cell.value = cell.value.replace('火蜀眼默卡', '火獨眼獸卡')
            cell.value = cell.value.replace('充满好奇心的学者幽需萌蟹', '充滿好奇心的學者幽靈萌獸')
            cell.value = cell.value.replace('古代径裔哥布林巫师萌蕙', '古代後裔哥布林巫師萌獸')
            cell.value = cell.value.replace('古代径裔哥布林刽士萌蕙','古代後裔哥布林劍士萌獸')
            cell.value = cell.value.replace('失去主人的看門狗萌默', '失去主人的看門狗萌獸')
            cell.value = cell.value.replace('巨大的咬食者萌默', '巨大的咬食者萌獸')
            cell.value = cell.value.replace('幼年扁普洛卡', '幼年烏普洛卡')
            cell.value = cell.value.replace('幼苗鳌萌默', '幼苗蠶萌獸')
            cell.value = cell.value.replace('幼黄蜀角狮卡', '幼黃獨角獅卡')
            cell.value = cell.value.replace('打倒像不出手的麻烦裂造者萌鲜', '打倒後不出手的麻煩製造者萌獸')
            cell.value = cell.value.replace('未知的鹿角免萌默', '未知的鹿角兔萌獸')
            cell.value = cell.value.replace('母狮萌默', '母獅萌獸')
            cell.value = cell.value.replace('生氧的鹿角免萌默', '生氣的鹿角兔萌獸')
            cell.value = cell.value.replace('生氧的辑鼠萌默', '生氣的溝鼠萌獸')
            cell.value = cell.value.replace('白雀萌默', '白雀萌獸')
            cell.value = cell.value.replace('白雪哨出岩萌默', '白雪噴出岩萌獸')
            cell.value = cell.value.replace('目里塔特萌默', '貝里塔特萌獸')
            cell.value = cell.value.replace('伊香薇逵(一般)萌默', '伊魯薇達(一般)萌獸')
            cell.value = cell.value.replace('伊香薇逵(狙擎)萌默', '伊魯薇達(狙擊)萌獸')
            cell.value = cell.value.replace('传鲁荷萌默', '後負荷萌獸')
            cell.value = cell.value.replace('光之轨行者卡', '光之執行者卡')
            cell.value = cell.value.replace('光之注视者卡', '光之注視者卡')
            cell.value = cell.value.replace('光璟缘水需萌默', '光環緣水靈萌獸')
            cell.value = cell.value.replace('冰之鹰萌默', '冰之鷹萌獸')
            cell.value = cell.value.replace('冰牛萌默', '冰牛萌獸')
            cell.value = cell.value.replace('冰莆卡', '冰龍卡')
            cell.value = cell.value.replace('刚玉警莆兵萌默', '剛玉警衛兵萌獸')
            cell.value = cell.value.replace('危除的石石森林昆毒萌蕙', '危險的石石森林昆蟲萌獸')
            cell.value = cell.value.replace('危除的香嘘免子萌默', '危險的香爐兔子萌獸')
            cell.value = cell.value.replace('危险的陶器免子萌默', '危險的陶器兔子萌獸')
            cell.value = cell.value.replace('危险的森林昆毒萌默', '危險的森林昆蟲萌獸')
            cell.value = cell.value.replace('吊克郎富的明萌默', '吊兒郎當的明萌獸')
            cell.value = cell.value.replace('地祭司萌默', '地祭司萌獸')
            cell.value = cell.value.replace('多利恩萌默', '多利恩萌獸')
            cell.value = cell.value.replace('奸笑的鳞芽木妖卡', '奸笑的發芽木妖卡')
            cell.value = cell.value.replace('守太初的某佃束西卡', '守護太初的某個東西卡')
            cell.value = cell.value.replace('安杰替斯萌默', '安杰魯斯萌獸')
            cell.value = cell.value.replace('安格洛楼器人A型萌默', '安格洛機器人 A型萌獸')
            cell.value = cell.value.replace('安格洛楼器人B型萌默', '安格洛機器人 B型萌獸')
            cell.value = cell.value.replace('安格洛楼器人C型萌默', '安格洛機器人 C型萌獸')
            cell.value = cell.value.replace('安普里欧卡', '安普里歐卡')
            cell.value = cell.value.replace('安萨斯迪温卡', '安薩斯迪溫卡')
            cell.value = cell.value.replace('托托萌默', '托托萌獸')
            cell.value = cell.value.replace('朱鹭萌默', '朱鷺萌獸')
            cell.value = cell.value.replace('死需史扁卡', '死靈史烏卡')
            cell.value = cell.value.replace('死需戴米安卡', '死靈戴米安卡')
            cell.value = cell.value.replace('灰鹰堆萌默', '灰塵堆萌獸')
            cell.value = cell.value.replace('米缝萌默', '米羅萌獸')
            cell.value = cell.value.replace('红赛王卡', '紅寶王卡')
            cell.value = cell.value.replace('红襄卡', '紅寶卡')
            cell.value = cell.value.replace('艾罪奈王萌默', '艾畢奈亞萌獸')
            cell.value = cell.value.replace('血刽隧昌萌默', '血劍隊員萌獸')
            cell.value = cell.value.replace('西米露谭子萌默', '西米露罈子萌獸')
            cell.value = cell.value.replace('西格鳍斯萌默', '西格諾斯萌獸')
            cell.value = cell.value.replace('乱七八糟雪吉拉萌默', '亂七八糟雪吉拉萌獸')
            cell.value = cell.value.replace('乱七八糟菇菇赛目萌默', '亂七八糟菇菇寶貝萌獸')
            cell.value = cell.value.replace('免子萌默', '兔子萌獸')
            cell.value = cell.value.replace('兵器楼器人typeB萌默', '兵器機器人typeB萌獸')
            cell.value = cell.value.replace('困鞋希拉卡', '困難希拉卡')
            cell.value = cell.value.replace('巡耀楼器人S卡', '巡邏機器人S卡')
            cell.value = cell.value.replace('库拉萌默', '庫拉萌獸')
            cell.value = cell.value.replace('戒翔里奥萌默', '飛翔里奧萌獸')
            cell.value = cell.value.replace('抓狂的家免萌默', '抓狂的家兔萌獸')
            cell.value = cell.value.replace('改造的作巢骨萌默', '改造的作業員萌獸')
            cell.value = cell.value.replace('杜那斯萌默', '杜那斯萌獸')
            cell.value = cell.value.replace('沉重的咬食者萌默', '沉重的咬食者萌獸')
            cell.value = cell.value.replace('沉默的值察者卡', '沉默的偵察者卡')
            cell.value = cell.value.replace('沙刃暴掠重萌默', '沙刃暴掠軍萌獸')
            cell.value = cell.value.replace('沙漠毒螺卡', '沙漠毒蠍卡')
            cell.value = cell.value.replace('沙鹰怪物萌默', '沙塵怪物萌獸')
            cell.value = cell.value.replace('纯白雪吉拉和企鹅或王卡', '純白雪吉拉和企鵝國王卡')
            cell.value = cell.value.replace('纯熟的德拉卡兹萌默', '純熟的德拉卡茲萌獸')
            cell.value = cell.value.replace('迅速的咬食者萌默', '迅速的咬食者萌獸')
            cell.value = cell.value.replace('那因哈特萌默', '那因哈特萌獸')
            cell.value = cell.value.replace('邪惠绵羊卡', '邪惡綿羊卡')
            cell.value = cell.value.replace('受逢的赛莲萌默', '受選的賽蓮萌獸')
            cell.value = cell.value.replace('受繁的狮子萌默', '受驚的獅子萌獸')
            cell.value = cell.value.replace('咂嘴咂舌的烈萌默', '咂嘴咂舌的烈萌獸')
            cell.value = cell.value.replace('夜晚守簧人萌默', '夜晚守護人萌獸')
            cell.value = cell.value.replace('姆呜水需萌窜', '姆嗚水靈萌寵')
            cell.value = cell.value.replace('届乞丐萌默', '風乞丐萌獸')
            cell.value = cell.value.replace('届之守潢露卡', '風之守護靈卡')
            cell.value = cell.value.replace('届祭司萌默', '風祭司萌獸')
            cell.value = cell.value.replace('怪物鼠萌默', '怪物鼠萌獸')
            cell.value = cell.value.replace('拉伊拉萌默', '拉伊拉萌獸')
            cell.value = cell.value.replace('沼墨烟泥怪物卡', '沼澤爛泥怪物卡')
            cell.value = cell.value.replace('波拉莆萌默', '波拉爾萌獸')
            cell.value = cell.value.replace('肯萌默', '肯萌獸')
            cell.value = cell.value.replace('花花蝶萌默', '花花蝶萌獸')
            cell.value = cell.value.replace('贪吃鬼卡', '貪吃鬼卡')
            cell.value = cell.value.replace('贪念的麻烦裂造者萌默', '貪念的麻煩製造者萌獸')
            cell.value = cell.value.replace('金图甲毒卡', '金屬甲蟲卡')
            cell.value = cell.value.replace('長相兑惠的赫萌默', '長相兇惡的赫萌獸')
            cell.value = cell.value.replace('長檐牛魔王卡', '長槍牛魔王卡')
            cell.value = cell.value.replace('阿瓦里蒂至', '阿瓦里蒂亞')
            cell.value = cell.value.replace('侵入的虚燕卡', '侵入的虛無卡')
            cell.value = cell.value.replace('南瓜幽需卡', '南瓜幽靈卡')
            cell.value = cell.value.replace('型甲鑫萌默', '聖甲蟲萌獸')
            cell.value = cell.value.replace('型骑士萌默', '聖騎士萌獸')
            cell.value = cell.value.replace('宫廷骑士萌默', '宮廷騎士萌獸')
            cell.value = cell.value.replace('帝顾接律萌默', '帝國護衛萌獸')
            cell.value = cell.value.replace('幽雯卡', '幽靈卡')
            cell.value = cell.value.replace('幽雯缎带肥肥萌默', '幽靈緞帶肥肥萌獸')
            cell.value = cell.value.replace('幽雯菇菇窝目萌默', '幽靈菇菇寶貝萌獸')
            cell.value = cell.value.replace('幽霸雪吉拉萌默', '幽靈雪吉拉萌獸')
            cell.value = cell.value.replace('扁鸦卡', '烏鴉卡')
            cell.value = cell.value.replace('柔道翁熊卡', '柔道貓熊卡')
            cell.value = cell.value.replace('段带肥肥卡', '緞帶肥肥卡')
            cell.value = cell.value.replace('洞窟地鼠萌默', '洞窟地鼠萌獸')
            cell.value = cell.value.replace('皇家簧律萌默', '皇家護衛萌獸')
            cell.value = cell.value.replace('突率型CQ57卡', '突擊型 CQ57卡')
            cell.value = cell.value.replace('美男法师萌默', '美男法師萌獸')
            cell.value = cell.value.replace('重雷長威甫_1皆段卡', '軍團長威爾_1階段卡')
            cell.value = cell.value.replace('重雷長威甫_2皆段卡', '軍團長威爾_2階段卡')
            cell.value = cell.value.replace('顺躁的疆属菇菇卡', '煩躁的殭屍菇菇卡')
            cell.value = cell.value.replace('哥布林巫师', '哥布林巫師')
            cell.value = cell.value.replace('哥布林刽士', '哥布林劍士')
            cell.value = cell.value.replace('哥布林盗贼', '哥布林盜賊')
            cell.value = cell.value.replace('哥布林歙缒兵', '哥布林鐵鎚兵')
            cell.value = cell.value.replace('哭泣的蒸菇菇卡', '哭泣的藍菇菇卡')
            cell.value = cell.value.replace('圆型自荷步檐萌默', '圓型負荷步槍萌獸')
            cell.value = cell.value.replace('弱化的幽囊削柬师萌默', '弱化的幽靈訓練師萌獸')
            cell.value = cell.value.replace('悚狱撤犬卡', '煉獄獵犬卡')
            cell.value = cell.value.replace('挫折的缘菇菇卡', '挫折的綠菇菇卡')
            cell.value = cell.value.replace('桃集练卡', '桃樂絲卡')
            cell.value = cell.value.replace('浪漫主羲者缘水需萌默', '浪漫主義者綠水靈萌獸')
            cell.value = cell.value.replace('浮化者卡', '淨化者卡')
            cell.value = cell.value.replace('海王蒂繇萌默', '海亞蒂絲萌獸')
            cell.value = cell.value.replace('海胎卡', '海膽卡')
            cell.value = cell.value.replace('海盗熊萌默', '海盜熊萌獸')
            cell.value = cell.value.replace('特蘭森迪温卡', '特蘭森迪溫卡')
            cell.value = cell.value.replace('疲傲的骑士萌默', '疲憊的騎士萌獸')
            cell.value = cell.value.replace('神型的古代刽选虎萌默', '神聖的古代劍齒虎萌獸')
            cell.value = cell.value.replace('神默卡', '神獸卡')
            cell.value = cell.value.replace('粉末蝴蝶萌默', '粉末蝴蝶萌獸')
            cell.value = cell.value.replace('翁缝将重卡', '翁羅將軍卡')
            cell.value = cell.value.replace('脍小的黑肥肥卡', '膽小的黑肥肥卡')
            cell.value = cell.value.replace('舰理大臣卡', '總理大臣卡')
            cell.value = cell.value.replace('豹萌默', '豹萌獸')
            cell.value = cell.value.replace('酒娜娜卡', '猶娜娜卡')
            cell.value = cell.value.replace('难减型T黑人钱B型萌默', '殲滅型T無人機B型萌獸')
            cell.value = cell.value.replace('难减型T黑人楼A型萌默', '殲滅型T無人機A型萌獸')
            cell.value = cell.value.replace('顿盔企鹅王卡', '頭盔企鵝王卡')
            cell.value = cell.value.replace('假面扁萌默', '假面鳥萌獸')
            cell.value = cell.value.replace('堕落魔族确化狼旗手卡', '墮落魔族強化狼旗手卡')
            cell.value = cell.value.replace('情怒的艾雨逵斯卡', '憤怒的艾爾達斯卡')
            cell.value = cell.value.replace('情怒的惠魔近莆兵萌默', '憤怒的惡魔近衛兵萌獸')
            cell.value = cell.value.replace('探索的福克斯里恩萌默', '探索的福克斯里恩萌獸')
            cell.value = cell.value.replace('敏捷的咬食者萌默', '敏捷的咬食者萌獸')
            cell.value = cell.value.replace('敏捷的洞窟地鼠萌默', '敏捷的洞窟地鼠萌獸')
            cell.value = cell.value.replace('旋国多利恩萌默', '旋風多利恩萌獸')
            cell.value = cell.value.replace('族畏巴萨克卡', '族長巴薩克卡')
            cell.value = cell.value.replace('梅杜莎萌默', '梅杜莎萌獸')
            cell.value = cell.value.replace('梅保德萌默', '梅傑德萌獸')
            cell.value = cell.value.replace('梦重人格的麻烦裂造者萌蟹', '雙重人格的麻煩製造者萌獸')
            cell.value = cell.value.replace('盗贼小姆勒卡', '盜賊小姆勒卡')
            cell.value = cell.value.replace('票泊者惹事份子卡', '漂泊者惹事份子卡')
            cell.value = cell.value.replace('绿水需卡', '綠水靈卡')
            cell.value = cell.value.replace('绿菇菇卡', '綠菇菇卡')
            cell.value = cell.value.replace('船昌克鲁卡', '船員克魯卡')
            cell.value = cell.value.replace('莉莉普雷王萌默', '莉莉普雷亞萌獸')
            cell.value = cell.value.replace('虚空的爪牙萌默', '虛空的爪牙萌獸')
            cell.value = cell.value.replace('被儒式港走的魔法师卡', '被儀式捲走的魔法師卡')
            cell.value = cell.value.replace('铜之肥肥卡', '鋼之肥肥卡')
            cell.value = cell.value.replace('铜载子弹毅手萌默', '鋼鐵子彈殺手萌獸')
            cell.value = cell.value.replace('铜载穆太卡', '鋼鐵穆太卡')
            cell.value = cell.value.replace('雪吉拉娃娃楼卡', '雪吉拉娃娃機卡')
            cell.value = cell.value.replace('雪免萌默', '雪免萌獸')
            cell.value = cell.value.replace('黄昏的爪牙萌默', '黃昏的爪牙萌獸')
            cell.value = cell.value.replace('黄金雪吉拉和企鹅或王卡', '黃金雪吉拉和企鵝國王卡')
            cell.value = cell.value.replace('傅脱的米诺陶洛斯', '傳說的米諾陶洛斯')
            cell.value = cell.value.replace('喜悦艾南逵斯卡', '喜悅艾爾達斯卡')
            cell.value = cell.value.replace('喵怪仙人(女)萌默', '喵怪仙人(女)萌獸')
            cell.value = cell.value.replace('喵怪仙人(男)萌默', '喵怪仙人(男)萌默')
            cell.value = cell.value.replace('奥芙赫班萌默', '奧芙赫班萌獸')
            cell.value = cell.value.replace('奥莉维王萌默', '奧莉維亞萌獸')
            cell.value = cell.value.replace('奥赛西翁卡', '奧賽西翁卡')
            cell.value = cell.value.replace('强化的重装(借萌默', '強化的重裝備萌獸')
            cell.value = cell.value.replace('强化钢载穆太卡', '強化鋼鐵穆太卡')
            cell.value = cell.value.replace('强化焉目立克β萌默', '強化馬貝立克β萌獸')
            cell.value = cell.value.replace('惠小丑磔克卡', '惡小丑傑克卡')
            cell.value = cell.value.replace('惠需卡', '惡靈卡')
            cell.value = cell.value.replace('惠魔水需萌默', '惡魔水靈萌獸')
            cell.value = cell.value.replace('惠魔绵羊卡', '惡魔綿羊卡')
            cell.value = cell.value.replace('惠魔搜索兵萌默', '惡魔搜索兵萌獸')
            cell.value = cell.value.replace('惠魔熊卡', '惡魔熊卡')
            cell.value = cell.value.replace('惠魔穆库香卡', '惡魔穆庫魯卡')
            cell.value = cell.value.replace('景影子骑士萌默', '影子騎士萌獸')
            cell.value = cell.value.replace('最终型熊涅涅卡', '最終型態涅涅卡')
            cell.value = cell.value.replace('最终型熊嘟嘟卡', '最終型態嘟嘟卡')
            cell.value = cell.value.replace('森林守簧人萌默', '森林守護人萌獸')
            cell.value = cell.value.replace('游魂卡', '遊魂卡')
            cell.value = cell.value.replace('痛魔法师卡力阿因卡', '瘋魔法師卡力阿因卡')
            cell.value = cell.value.replace('缘海焉卡', '綠海馬卡')
            cell.value = cell.value.replace('菇菇赛目卡', '菇菇寶貝卡')
            cell.value = cell.value.replace('装可卡', '裘可卡')
            cell.value = cell.value.replace('谢娜蜀士伽玛版本萌默', '謝娜戰士伽瑪版本萌獸')
            cell.value = cell.value.replace('谢娜蜀士阿葡法版本萌蕙', '謝娜戰士阿爾法版本萌獸')
            cell.value = cell.value.replace('谢娜蜀士蓓塔版本萌默', '謝娜戰士蓓塔版本萌獸')
            cell.value = cell.value.replace('超望的刀刃卡', '絕望的刀刃卡')
            cell.value = cell.value.replace('量随的麻烦裂造者萌默', '單戀的麻煩製造者萌獸')
            cell.value = cell.value.replace('锂磺穆太卡', '鋰礦穆太卡')
            cell.value = cell.value.replace('雲之守潢露卡', '雲之守護靈卡')
            cell.value = cell.value.replace('鱿鱼的麻烦裂造者萌默', '魷魚的麻煩製造者萌獸')
            cell.value = cell.value.replace('黑名魔默萌默', '無名魔獸萌獸')
            cell.value = cell.value.replace('黑色太陨步兵萌默', '黑色太陽步兵萌獸')
            cell.value = cell.value.replace('黑色太隔弓兵萌默', '黑色太陽弓兵萌獸')
            cell.value = cell.value.replace('黑色太隔炸强兵萌默', '黑色太陽炸彈兵萌獸')
            cell.value = cell.value.replace('黑色太隔魔法兵萌默', '黑色太陽魔法兵萌獸')
            cell.value = cell.value.replace('黑能卡', '黑龍卡')
            cell.value = cell.value.replace('黑骑士魔凯丁卡', '黑騎士魔凱丁卡')
            cell.value = cell.value.replace('黑暗莱西卡', '黑暗萊西卡')
            cell.value = cell.value.replace('黑漆漆的嫩窦萌默', '黑漆漆的嫩寶萌獸')
            cell.value = cell.value.replace('黑蝴蛛卡', '黑蜘蛛卡')
            cell.value = cell.value.replace('摇控晃晃的际纹葫蔗萌蕙', '搖搖晃晃的條紋葫蘆萌獸')
            cell.value = cell.value.replace('新复古红螃壁萌默', '新復古紅螃蟹萌獸')
            cell.value = cell.value.replace('新复古巫婆萌默', '新復古巫婆萌獸')
            cell.value = cell.value.replace('新复古钢之肥肥萌默', '新復古鋼之肥肥萌獸')
            cell.value = cell.value.replace('新复古理属猴王萌默', '新復古殭屍猴王萌獸')
            cell.value = cell.value.replace('新复古猴子萌默', '新復古猴子萌獸')
            cell.value = cell.value.replace('新复古缘水盂萌默', '新復古綠水靈萌獸')
            cell.value = cell.value.replace('暗之轨行者卡', '暗之執行者卡')
            cell.value = cell.value.replace('暗黑三角能卡', '暗黑三角龍卡')
            cell.value = cell.value.replace('暗蜀眼默卡', '暗獨眼獸卡')
            cell.value = cell.value.replace('莱西卡', '萊西卡')
            cell.value = cell.value.replace('萼中的露希妲卡', '夢中的露希妲卡')
            cell.value = cell.value.replace('落花免萌默', '落花兔萌獸')
            cell.value = cell.value.replace('葛雷雪吉拉和企鹅或王卡', '葛雷雪吉拉和企鵝國王卡')
            cell.value = cell.value.replace('葵色龍', '變色龍')
            cell.value = cell.value.replace('葵身娃娃栈卡', '變身娃娃機卡')
            cell.value = cell.value.replace('蜀角尼莫卡', '獨角尼莫卡')
            cell.value = cell.value.replace('路澄卡', '路燈卡')
            cell.value = cell.value.replace('遇离警莆兵萌默', '瑪瑙警衛兵萌獸')
            cell.value = cell.value.replace('遇蒙萌默', '瑪蒙萌獸')
            cell.value = cell.value.replace('遗古妖精萌默', '遠古妖精萌獸')
            cell.value = cell.value.replace('隔光精需卡', '陽光精靈卡')
            cell.value = cell.value.replace('雷普暗毅者萌默', '雷普暗殺者萌獸')
            cell.value = cell.value.replace('雷普楼甲兵萌默', '雷普機甲兵萌獸')
            cell.value = cell.value.replace('鼠田重上级值察兵卡', '織田軍上級偵察兵卡')
            cell.value = cell.value.replace('鼠田重武士卡', '織田軍武士卡')
            cell.value = cell.value.replace('鼠田重步兵卡', '織田軍步兵卡')
            cell.value = cell.value.replace('鼠田重值察兵卡', '織田軍偵察兵卡')
            cell.value = cell.value.replace('鼠田重除隔师卡', '織田軍陰陽師卡')
            cell.value = cell.value.replace('墓威菇菇卡', '殭屍菇菇卡')
            cell.value = cell.value.replace('嫩赛卡', '嫩寶卡')
            cell.value = cell.value.replace('熙鱼克洛克卡', '鱷魚克洛克卡')
            cell.value = cell.value.replace('疑赫萌默', '羅赫萌獸')
            cell.value = cell.value.replace('精水的谭子萌默', '積水的罈子萌獸')
            cell.value = cell.value.replace('精妖萌默', '樹妖萌獸')
            cell.value = cell.value.replace('蒲川一益卡', '瀧川一益卡')
            cell.value = cell.value.replace('蒸水需卡', '藍水靈卡')
            cell.value = cell.value.replace('蒸色泉之精需萌默', '藍色泉之精靈萌獸')
            cell.value = cell.value.replace('蒸色缎带肥肥卡', '藍色緞帶肥肥卡')
            cell.value = cell.value.replace('蒸色燮角能卡', '藍色雙角龍卡')
            cell.value = cell.value.replace('蒸色蘑菇王萌默', '藍色蘑菇王萌獸')
            cell.value = cell.value.replace('蒸菇菇卡', '藍菇菇卡')
            cell.value = cell.value.replace('蒸赛卡', '藍寶卡')
            cell.value = cell.value.replace('辖水贼卡', '竊水賊卡')
            cell.value = cell.value.replace('遣跻魔螺卡', '遺跡魔蠍卡')
            cell.value = cell.value.replace('鲜奶油雷卡', '鮮奶油團卡')
            cell.value = cell.value.replace('鼻找险景泊的圆球萌默', '尋找陰影的圓球萌獸')
            cell.value = cell.value.replace('鼻找隔光的圆球萌默', '尋找陽光的圓球萌獸')
            cell.value = cell.value.replace('德奈波拉萌默', '德奈波拉萌獸')
            cell.value = cell.value.replace('暴届卡', '暴風卡')
            cell.value = cell.value.replace('潘斯萌默', '潘斯萌獸')
            cell.value = cell.value.replace('澜弄矩卡', '彌弄矩卡')
            cell.value = cell.value.replace('熟棘的德拉卡兹萌默', '熟練的德拉卡茲萌獸')
            cell.value = cell.value.replace('稻草人萌默', '稻草人萌獸')
            cell.value = cell.value.replace('蝙虫雷卡', '蝙蝠卡')
            cell.value = cell.value.replace('觐律隧企鹅王卡', '親衛隊企鵝王卡')
            cell.value = cell.value.replace('鲤翁卡', '狸貓卡')
            cell.value = cell.value.replace('橡木甲毒卡', '橡木甲蟲卡')
            cell.value = cell.value.replace('燃赎的情怒卡', '燃燒的情怒卡')
            cell.value = cell.value.replace('蕙型衡南瓜橙卡', '萬聖節南瓜燈卡')
            cell.value = cell.value.replace('磷芽木妖卡', '發芽木妖卡')
            cell.value = cell.value.replace('鳅免子', '戰鬥兔子')
            cell.value = cell.value.replace('鳅型T黑人钱B型萌默', '戰鬥型T無人機B型萌獸')
            cell.value = cell.value.replace('鳅型T黑人楼A型萌默', '戰鬥型T無人機A型萌獸')
            cell.value = cell.value.replace('黏呼呼莱子管蝇萌默', '黏呼呼葉子蒼蠅萌獸')
            cell.value = cell.value.replace('黏站踢踢莱子管蝇萌默', '黏踢踢葉子蒼蠅萌獸')
            cell.value = cell.value.replace('鹰眼卡', '鷹眼卡')
            cell.value = cell.value.replace('警律楼器人C1-49萌默', '警衛機器人 C1-49萌獸')
            cell.value = cell.value.replace('魔王幽需卡', '魔王幽靈卡')
            cell.value = cell.value.replace('魔莆卡', '魔龍卡')
            cell.value = cell.value.replace('魔靳隧骨萌默', '魔斬隊員萌獸')
            cell.value = cell.value.replace('小精需游侠卡', '小精靈遊俠卡')
            cell.value = cell.value.replace('小毅人卡', '小殺人卡')
            cell.value = cell.value.replace('幼蒸蜀角狮卡', '幼藍獨角獅卡')
            cell.value = cell.value.replace('巡耀楼器人卡', '巡邏機器人卡')
            
            if cell.value == "攻擊力" or cell.value == "擊力" or cell.value == "果":
                # 取得儲存格的位置
                col_index = cell.column
                row_index = cell.row
                
                # 取得位置
                a_index = str(cell.value).index(cell.value)
                
                # 將cell插入左邊欄位的最後面
                if ws.cell(row=row_index, column=col_index-1).value != None:
                    ws.cell(row=row_index, column=col_index-1).value = ws.cell(row=row_index, column=col_index-1).value + cell.value
                else:
                    ws.cell(row=row_index, column=col_index-1).value = cell.value
                # ws.cell(row=row_index, column=col_index).value = cell.value
                
                # 清空cell所在的儲存格
                cell.value = None
                
                # 將cell右邊的欄位往左移動一格
                for i in range(col_index+1, ws.max_column+1):
                    ws.cell(row=row_index, column=i-1).value = ws.cell(row=row_index, column=i).value
                    ws.cell(row=row_index, column=i).value = None
            
        # 計算目前的進度完成率
        progress_value = get_progress_value(row, col)
        progress_percent = (progress_value - start_value) / (end_value - start_value) * 100
        # 顯示目前進度完成率
        print(f'目前進度：{progress_percent:.2f}%')
        

# 儲存檔案
wb.save('maple-auto-auction-after.xlsx')
